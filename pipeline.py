"""
ProcessOptimizer KZ — Raw Excel Ingestion Pipeline
====================================================
Reads raw XHQ historian Excel files, merges them into a single hourly CSV.

Output feeds the Airflow ingest_etl_dag (source_path = data/processed/cleaned_dataset.csv).

Run once per data drop:
    python pipeline.py

Output:
    data/processed/cleaned_dataset.csv
"""

import os
from datetime import datetime

import numpy as np
import openpyxl
import pandas as pd

# ── CONFIG ────────────────────────────────────────────────────────────────────
RAW_DATA_DIR = "data/raw/"
OUTPUT_DIR   = "data/processed/"
OUTPUT_FILE  = "cleaned_dataset.csv"

# ── SENSOR DEFINITIONS ────────────────────────────────────────────────────────
# Simple sensors: one value + timestamp per row.
SENSORS_SIMPLE = {
    "TE301020": {
        "file": "DiffPressTransm_TT301020.xlsx",
        "unit": "degC",
        "desc": "DEA amine temperature",
        "plant": "D304",
    },
    "PDT31008": {
        "file": "DiffPressTransm_PDT31008.xlsx",
        "unit": "mbar",
        "desc": "Pressure differential D304",
        "plant": "D304",
    },
    "PDT31001": {
        "file": "DiffPressTransm_PDT31001.xlsx",
        "unit": "mbar",
        "desc": "Pressure differential D301",
        "plant": "D301",
    },
    "PDT31007": {
        "file": "DiffPressTransm_PDT31007_archive.xlsx",
        "unit": "mbar",
        "desc": "Pressure differential D304 archive",
        "plant": "D304",
    },
    "FQ31050": {
        "file": "DiffPressTransm_TT31050.xlsx",
        "unit": "m3/h",
        "desc": "Steam flow D304",
        "plant": "D304",
    },
    "LT301031": {
        "file": "DiffPressTransm_LT301031.xlsx",
        "unit": "%",
        "desc": "Level D304 (volatile, high variability)",
        "plant": "D304",
    },
}

# Controller sensors: PV (process value) + OP (output) + SP (setpoint).
# Column layout in Excel: col[0]=PV value, col[1]=PV ts | col[4]=OP value,
# col[5]=OP ts | col[8]=SP value, col[9]=SP ts.
SENSORS_CTRL = {
    "LIC31012": {
        "file": "DiffPressTransm_Lt31012.xlsx",
        "unit": "%",
        "desc": "Level controller D304 (SP=80% constant)",
        "plant": "D304",
        "sp_fixed": 80.0,
    },
    "LIC31002": {
        "file": "DiffPressTransm_Lt31002.xlsx",
        "unit": "%",
        "desc": "Level controller D301 (SP changes: 65→70→92→70)",
        "plant": "D301",
        "sp_fixed": None,
    },
    "FIC31011": {
        "file": "DiffPressTransm_LT31011.xlsx",
        "unit": "m3/h",
        "desc": "Reflux flow controller D304 (SP dynamic)",
        "plant": "D304",
        "sp_fixed": None,
    },
}


# ── LOADERS ───────────────────────────────────────────────────────────────────

def load_simple(sensor_id: str, meta: dict, data_dir: str) -> pd.DataFrame:
    """Load a simple sensor Excel file (value, timestamp)."""
    path = os.path.join(data_dir, meta["file"])
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    data = []
    for row in rows[2:]:   # skip header + blank row
        if row[0] is None or row[1] is None:
            continue
        if isinstance(row[0], (int, float)) and isinstance(row[1], datetime):
            data.append({"timestamp": row[1], sensor_id: float(row[0])})

    df = pd.DataFrame(data)
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    return df.set_index("timestamp").sort_index()


def load_controller(sensor_id: str, meta: dict, data_dir: str) -> pd.DataFrame:
    """Load a controller Excel file (PV + OP + SP in wide layout)."""
    path = os.path.join(data_dir, meta["file"])
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    pv_data, op_data, sp_data = [], [], []
    for row in rows[2:]:
        if len(row) < 3:
            continue
        if row[0] is not None and isinstance(row[0], (int, float)) and isinstance(row[1], datetime):
            pv_data.append({"timestamp": row[1], f"{sensor_id}_PV": float(row[0])})
        if len(row) > 6 and row[4] is not None and isinstance(row[4], (int, float)) and isinstance(row[5], datetime):
            op_data.append({"timestamp": row[5], f"{sensor_id}_OP": float(row[4])})
        if len(row) > 10 and row[8] is not None and isinstance(row[8], (int, float)) and isinstance(row[9], datetime):
            sp_data.append({"timestamp": row[9], f"{sensor_id}_SP": float(row[8])})

    dfs = []
    for data in [pv_data, op_data, sp_data]:
        if data:
            d = pd.DataFrame(data).set_index("timestamp").sort_index()
            d.index = pd.to_datetime(d.index)
            dfs.append(d)

    return pd.concat(dfs, axis=1, sort=True) if dfs else pd.DataFrame()


# ── MAIN PIPELINE ─────────────────────────────────────────────────────────────

def run_pipeline(
    data_dir: str = RAW_DATA_DIR,
    output_dir: str = OUTPUT_DIR,
    resample_freq: str = "1h",
    ffill_limit: int = 6,
) -> pd.DataFrame:
    """
    Full ingestion pipeline:
      1. Load all raw Excel files
      2. Resample to a uniform hourly grid (median aggregation)
      3. Forward-fill gaps (event-driven sensors log only on value change)
      4. Engineer features: deviations, time columns
      5. Interpolate + median fallback for remaining NaN
      6. Save cleaned_dataset.csv

    Returns cleaned, ML-ready DataFrame.
    """
    os.makedirs(output_dir, exist_ok=True)

    # ── 1. Load ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("STEP 1 — Loading raw sensor files")
    print("=" * 60)
    all_dfs = []

    for sid, meta in SENSORS_SIMPLE.items():
        df = load_simple(sid, meta, data_dir)
        print(f"  {sid:12s} | {len(df):5d} raw records | {meta['unit']:6s} | {meta['plant']}")
        all_dfs.append(df.resample(resample_freq).median())

    for sid, meta in SENSORS_CTRL.items():
        df = load_controller(sid, meta, data_dir)
        print(f"  {sid:12s} | {len(df):5d} raw records | {meta['unit']:6s} | {meta['plant']} (PV+OP+SP)")
        all_dfs.append(df.resample(resample_freq).median())

    # ── 2. Merge ──────────────────────────────────────────────────────────────
    print("\nSTEP 2 — Merging into single DataFrame")
    merged = pd.concat(all_dfs, axis=1, sort=True)
    print(f"  Shape: {merged.shape}  |  {merged.index.min()} → {merged.index.max()}")
    print(f"  Missing before fill: {merged.isnull().sum().sum()} cells "
          f"({merged.isnull().sum().sum() / merged.size * 100:.1f}%)")

    # ── 3. Forward-fill ───────────────────────────────────────────────────────
    # Sensors are event-driven — they only log when value changes.
    # After hourly resample, quiet hours appear as NaN.
    print(f"\nSTEP 3 — Forward-fill (limit={ffill_limit}h)")
    filled = merged.ffill(limit=ffill_limit)
    print(f"  Missing after ffill: {filled.isnull().sum().sum()} cells")

    # ── 4. Feature engineering ────────────────────────────────────────────────
    print("\nSTEP 4 — Feature engineering")

    # LIC31012 SP is physically constant at 80% — restore before deviation.
    if "LIC31012_SP" in filled.columns:
        filled["LIC31012_SP"] = filled["LIC31012_SP"].fillna(80.0)

    # Controller deviations (SP − PV convention, consistent with 02_models.ipynb).
    for ctrl in ["LIC31012", "LIC31002", "FIC31011"]:
        pv_col  = f"{ctrl}_PV"
        sp_col  = f"{ctrl}_SP"
        dev_col = f"{ctrl}_deviation"
        if pv_col in filled.columns and sp_col in filled.columns:
            filled[sp_col]  = filled[sp_col].ffill()
            filled[dev_col] = filled[sp_col] - filled[pv_col]
            dev = filled[dev_col].dropna()
            print(f"  {dev_col}: mean={dev.mean():.2f}  std={dev.std():.2f}  max_abs={dev.abs().max():.2f}")

    # Time features (required by POINT_FEATURES in train_job.py).
    filled["hour_of_day"] = filled.index.hour
    filled["day_of_week"] = filled.index.dayofweek
    print("  Added hour_of_day, day_of_week")

    # ── 5. Final imputation ───────────────────────────────────────────────────
    print("\nSTEP 5 — Final imputation")
    cleaned = filled.interpolate(method="time", limit=3)
    for col in cleaned.columns:
        if cleaned[col].isnull().any():
            med = cleaned[col].median()
            cleaned[col] = cleaned[col].fillna(med)
    print(f"  Final NaN count: {cleaned.isnull().sum().sum()}")

    # ── 6. Save ───────────────────────────────────────────────────────────────
    out_path = os.path.join(output_dir, OUTPUT_FILE)
    cleaned.index.name = "timestamp"
    cleaned.to_csv(out_path)

    print("\n" + "=" * 60)
    print("PIPELINE COMPLETE")
    print("=" * 60)
    print(f"  Output : {out_path}")
    print(f"  Shape  : {cleaned.shape}  ({cleaned.shape[0]} hours × {cleaned.shape[1]} features)")
    print(f"  Period : {cleaned.index.min()} → {cleaned.index.max()}")
    return cleaned


def print_quality_report(df: pd.DataFrame) -> None:
    """Print a quick data quality summary and known anomalous periods."""
    print("\n" + "=" * 60)
    print("DATA QUALITY REPORT")
    print("=" * 60)
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    stats = df[numeric_cols].agg(["min", "max", "mean", "std", "median"])
    print(stats.T.round(3).to_string())
    print("\nKnown anomalous periods (from visual inspection):")
    print("  25-26 Dec 2014  — FIC31011 spike (PV=448 m3/h, Z=8.0)")
    print("  28-30 Dec 2014  — PDT31008 high pressure (300+ mbar)")
    print("  02-03 Jan 2015  — MAJOR: LIC31002 dev=+20.9%, LT301031 swings 8-94%")
    print("  10-11 Jan 2015  — PDT31008 maximum (378 mbar)")
    print("  12-13 Jan 2015  — MAJOR: LT301031 swings, FQ31050 drops")
    print("  18 Jan 2015     — LIC31002 dev=+12%")


if __name__ == "__main__":
    cleaned = run_pipeline(
        data_dir="data/raw/",
        output_dir="data/processed/",
        resample_freq="1h",
        ffill_limit=6,
    )
    print_quality_report(cleaned)
